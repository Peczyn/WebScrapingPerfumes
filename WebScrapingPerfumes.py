#app worked back then but site changed so it doesn't now
#had no need to repair it so left it like that
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook


def main():
    wb = load_workbook('Perfumy.xlsx') #nazwa pliku xlsx który będziemy wczytywać 
    ws = wb.active #wb workbook ws worksheet
    i=2 #licznik rzędu w excelu

    
    for page_counter in range(1): #range is number of pages we want to visit
        
        #one page is 25 products and the 9 at the end of a link declares sorting by bestsellers
        url = f"https://www.notino.pl/perfumy/?f={page_counter}-9-55544"

        #send request to url and save reply in doc
        result = requests.get(url).text
        doc = BeautifulSoup(result, "html.parser")
       

        #now we iterate through all products
        for product_counter in range(24):

            #for each product we need to get their code
            products_on_page = doc.find_all('a',class_="sc-iOeugr iiaXZj")
            


            #find link to product and save link in excel file
            try:
                page_adress = str(products_on_page[product_counter]).split('href="')[1].split('"')[0]
                ws[f'G{i}'] = page_adress
                print(i-1,page_adress) #shows that program is working
            except IndexError:
                print("INDEXERR1")
                continue
            
            
        
            #send request for each product website
            url = f"https://www.notino.pl{page_adress}"
            result = requests.get(url).text
            doc1 = BeautifulSoup(result, "html.parser")
            
            #get the name of the product
            name = doc1.find('span',class_="sc-3sotvb-4 kSRNEJ").text
            
            #get the producent of product
            producent = doc1.find('a',class_="sc-3sotvb-2 iYvTNX").text
            
            #get the different smells of the product
            smells= doc1.find_all('td',class_="sc-1eu1dd2-9 jCBSyO")
           
            try:
                smell_head = smells[0].text
                ws[f'C{i}'] = smell_head
            except IndexError:
                ws[f'C{i}'] = "---"
            
            try:    
                smell_heart = smells[1].text
                ws[f'D{i}'] = smell_heart
            except IndexError:
                ws[f'D{i}'] = "---"
           
            try:    
                smell_base = smells[2].text
                ws[f'E{i}'] = smell_base
            except IndexError:
                ws[f'E{i}'] = "---"
            
            try:
                smell_group = zapachy[3].text
                ws[f'F{i}'] = smell_group
            except IndexError:
                ws[f'F{i}'] = "---"
            
            ws[f'A{i}'] = name
            ws[f'B{i}'] = producent
   
            i+=1
        wb.save('Perfumyv.xlsx') #we need to save our file
    print("DONE")
    
main()   